import openpyxl
import os
import re
from datetime import datetime
from sqlalchemy.orm import Session
from backend.database import (
    Faculty, Subject, Section, Batch, Room, TimeSlot, Assignment, LoadDistribution, User, hash_password
)

def clean_text(val):
    if val is None:
        return ""
    return str(val).strip()

def clean_id(name):
    if not name:
        return ""
    name = str(name).lower()
    for prefix in ["dr (mrs)", "dr (ms)", "dr.", "mr.", "ms.", "mrs.", "prof."]:
        if name.startswith(prefix):
            name = name[len(prefix):]
    name = re.sub(r'[^a-z0-9\s]', '', name)
    return "_".join(name.split())

def get_cell_value(sheet, r, c):
    val = sheet.cell(row=r, column=c).value
    # Check if cell is in a merged range
    for merged_range in sheet.merged_cells.ranges:
        if r >= merged_range.min_row and r <= merged_range.max_row:
            if c >= merged_range.min_col and c <= merged_range.max_col:
                return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return val

def extract_date_from_string(text):
    if not text:
        return None
    match = re.search(r'(\d{1,2})[.-](\d{1,2})[.-](\d{2,4})', text)
    if match:
        d, m, y = match.group(1), match.group(2), match.group(3)
        if len(y) == 2:
            y = "20" + y
        try:
            return f"{y}-{int(m):02d}-{int(d):02d}"
        except ValueError:
            return None
    return None

def extract_faculty_legend(text):
    if not text:
        return []
    text = clean_text(text)
    parts = re.split(r'/|\+|&|\band\b', text)
    matches = []
    for part in parts:
        part = part.strip()
        match = re.search(r'([^(]+)\(([^)]+)\)', part)
        if match:
            full_name = match.group(1).strip()
            initials = match.group(2).strip()
            matches.append((initials, full_name))
        else:
            if len(part) > 2:
                matches.append(("", part))
    return matches

def parse_legends(wb):
    faculty_map = {}
    subject_map = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        legend_row = None
        for r in range(1, sheet.max_row + 1):
            row_vals = [clean_text(get_cell_value(sheet, r, c)) for c in range(1, 10)]
            if any("sub. code" in x.lower() or "faculty" in x.lower() for x in row_vals):
                legend_row = r
                break
                
        if legend_row:
            for r in range(legend_row + 1, sheet.max_row + 1):
                code = clean_text(get_cell_value(sheet, r, 3))
                subj = clean_text(get_cell_value(sheet, r, 4))
                fac = clean_text(get_cell_value(sheet, r, 5))
                
                if code and subj:
                    subject_map[code] = subj
                if fac:
                    fac_extracted = extract_faculty_legend(fac)
                    for initials, full_name in fac_extracted:
                        if initials:
                            faculty_map[initials.upper()] = full_name
                            
    return faculty_map, subject_map

def find_grid2_start_col(sheet):
    for col in range(10, sheet.max_column + 1):
        for r in range(7, 23):
            val = get_cell_value(sheet, r, col)
            if val is not None and str(val).strip() != "":
                return col
    return None

def parse_grid_cell(room_val, subject_val, faculty_val):
    if not subject_val or str(subject_val).strip() == "" or str(subject_val).lower() in ["free", "lunch", "sport", "library", "sports"]:
        return []
    
    room_str = clean_text(room_val)
    sub_str = clean_text(subject_val)
    fac_str = clean_text(faculty_val)
    
    room_parts = [r.strip() for r in room_str.split("/")] if "/" in room_str else [room_str]
    sub_parts = [s.strip() for s in sub_str.split("/")] if "/" in sub_str else [sub_str]
    fac_parts = [f.strip() for f in fac_str.split("/")] if "/" in fac_str else [fac_str]
    
    if len(sub_parts) == 2 and len(fac_parts) == 2:
        r_parts = room_parts if len(room_parts) == 2 else [room_str, room_str]
        res = []
        for i in range(2):
            s_part = sub_parts[i]
            f_part = fac_parts[i]
            r_part = r_parts[i]
            
            batch = None
            if "B1" in s_part or "b1" in s_part.lower():
                batch = "B1"
            elif "B2" in s_part or "b2" in s_part.lower():
                batch = "B2"
                
            clean_sub = re.sub(r'\b(B1|B2|b1|b2)\b', '', s_part).strip()
            clean_sub = " ".join(clean_sub.split())
            res.append({
                "room": r_part,
                "subject": clean_sub,
                "faculty": f_part,
                "batch": batch
            })
        return res
        
    if len(fac_parts) == 2 and ("&" in sub_str or "," in sub_str or "and" in sub_str.lower()) and ("B1" in sub_str and "B2" in sub_str):
        clean_sub = re.sub(r'\b(B1|B2|b1|b2|&|,|and)\b', '', sub_str).strip()
        clean_sub = " ".join(clean_sub.split())
        return [
            {"room": room_str, "subject": clean_sub, "faculty": fac_parts[0], "batch": "B1"},
            {"room": room_str, "subject": clean_sub, "faculty": fac_parts[1], "batch": "B2"}
        ]
        
    batch = None
    if "B1" in sub_str or "b1" in sub_str.lower():
        batch = "B1"
    elif "B2" in sub_str or "b2" in sub_str.lower():
        batch = "B2"
        
    clean_sub = re.sub(r'\b(B1|B2|b1|b2)\b', '', sub_str).strip()
    clean_sub = " ".join(clean_sub.split())
    
    return [{
        "room": room_str,
        "subject": clean_sub,
        "faculty": fac_str,
        "batch": batch
    }]

def parse_load_distribution(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb["Final Subject Distribution "]
    
    records = []
    curr_faculty = ""
    curr_semester = ""
    
    for r in range(6, sheet.max_row + 1):
        f_name = clean_text(get_cell_value(sheet, r, 2))
        sem = clean_text(get_cell_value(sheet, r, 3))
        sec = clean_text(get_cell_value(sheet, r, 4))
        sub = clean_text(get_cell_value(sheet, r, 5))
        th = get_cell_value(sheet, r, 6)
        pr = get_cell_value(sheet, r, 7)
        tot = get_cell_value(sheet, r, 10)
        
        if f_name:
            curr_faculty = f_name
        if sem:
            curr_semester = sem
            
        if not sec or not sub:
            continue
            
        try:
            th_val = float(th) if th is not None else 0.0
            pr_val = float(pr) if pr is not None else 0.0
            tot_val = float(tot) if tot is not None else (th_val + pr_val)
        except ValueError:
            th_val, pr_val, tot_val = 0.0, 0.0, 0.0
            
        records.append({
            "faculty_name": curr_faculty,
            "semester": curr_semester,
            "section_name": sec,
            "subject_name": sub,
            "theory_hours": th_val,
            "practical_hours": pr_val,
            "total_hours": tot_val
        })
    return records

def parse_class_wise_timetables(path, year_label, global_faculty_mapping):
    wb = openpyxl.load_workbook(path, data_only=True)
    assignments = []
    
    all_dates = set()
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for r in range(1, 15):
            for c in range(1, sheet.max_column + 1):
                val = get_cell_value(sheet, r, c)
                if val and "effective from" in str(val).lower():
                    d = extract_date_from_string(str(val))
                    if d:
                        all_dates.add(d)
                        
    date_list = sorted(list(all_dates))
    if not date_list:
        date_list = ["2024-09-30", "2024-11-18"]
    elif len(date_list) == 1:
        date_list.append("2024-11-18" if date_list[0] != "2024-11-18" else "2024-09-30")
        
    date_list = sorted(date_list)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        section_name = sheet_name.strip()
        dept = "CS"
        if "it" in section_name.lower():
            dept = "IT"
        elif "aiml" in section_name.lower():
            dept = "AIML"
        elif "iot" in section_name.lower():
            dept = "IoT"
        elif "ds" in section_name.lower():
            dept = "DS"
            
        section_id = f"{section_name.replace(' ', '_')}_{year_label}"
        
        grid1_date = None
        for r in range(1, 15):
            for c in range(1, sheet.max_column + 1):
                val = get_cell_value(sheet, r, c)
                if val and "effective from" in str(val).lower():
                    grid1_date = extract_date_from_string(str(val))
                    break
            if grid1_date:
                break
                
        if not grid1_date:
            grid1_date = date_list[0]
            
        grid2_date = date_list[1] if grid1_date == date_list[0] else date_list[0]
        
        day_rows = {}
        for r in range(1, sheet.max_row + 1):
            val = clean_text(sheet.cell(row=r, column=1).value).upper()
            if val in ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"]:
                day_rows[val] = r
                
        # Determine if Grid 2 is shifted vertically
        grid2_col = find_grid2_start_col(sheet)
        is_grid2_shifted = False
        if grid2_col:
            # Check if Row 7 has data in columns grid2_col to grid2_col + 7
            for c in range(grid2_col, grid2_col + 8):
                val = get_cell_value(sheet, 7, c)
                if val is not None and str(val).strip() != "":
                    is_grid2_shifted = True
                    break

        # Parse Grid 1 (cols 2-9)
        for day, r_day in day_rows.items():
            for p in range(1, 9):
                col = p + 1
                room_val = get_cell_value(sheet, r_day, col)
                subject_val = get_cell_value(sheet, r_day + 1, col)
                faculty_val = get_cell_value(sheet, r_day + 2, col)
                
                # Check for default classroom fallback
                if room_val == "L" or not room_val:
                    # Find default classroom for this section
                    # We can use the classroom info from row 5
                    room_val = f"Room_{section_name}"
                
                parsed_slots = parse_grid_cell(room_val, subject_val, faculty_val)
                for slot in parsed_slots:
                    fac_initials = slot["faculty"].strip().upper()
                    fac_name = global_faculty_mapping.get(fac_initials, fac_initials)
                    fac_id = clean_id(fac_name) if fac_name else "unknown"
                    
                    assignments.append({
                        "faculty_id": fac_id,
                        "faculty_name": fac_name or fac_initials,
                        "faculty_initials": fac_initials,
                        "subject_code": slot["subject"],
                        "section_id": section_id,
                        "section_name": section_name,
                        "batch_label": slot["batch"],
                        "room_id": slot["room"],
                        "day": day.capitalize(),
                        "period": p,
                        "effective_from": grid1_date,
                        "effective_to": grid2_date if grid1_date < grid2_date and grid2_col else None,
                        "source": "class_wise"
                    })
                    
        # Parse Grid 2 (if present)
        if grid2_col:
            for day, r_day in day_rows.items():
                r_start = r_day - 1 if is_grid2_shifted else r_day
                for p in range(1, 9):
                    col = grid2_col + p - 1
                    room_val = get_cell_value(sheet, r_start, col)
                    subject_val = get_cell_value(sheet, r_start + 1, col)
                    faculty_val = get_cell_value(sheet, r_start + 2, col)
                    
                    if room_val == "L" or not room_val:
                        room_val = f"Room_{section_name}"
                        
                    parsed_slots = parse_grid_cell(room_val, subject_val, faculty_val)
                    for slot in parsed_slots:
                        fac_initials = slot["faculty"].strip().upper()
                        fac_name = global_faculty_mapping.get(fac_initials, fac_initials)
                        fac_id = clean_id(fac_name) if fac_name else "unknown"
                        
                        assignments.append({
                            "faculty_id": fac_id,
                            "faculty_name": fac_name or fac_initials,
                            "faculty_initials": fac_initials,
                            "subject_code": slot["subject"],
                            "section_id": section_id,
                            "section_name": section_name,
                            "batch_label": slot["batch"],
                            "room_id": slot["room"],
                            "day": day.capitalize(),
                            "period": p,
                            "effective_from": grid2_date,
                            "effective_to": grid1_date if grid2_date < grid1_date else None,
                            "source": "class_wise"
                        })
                        
    return assignments

def parse_individual_faculty_wise(path, global_faculty_mapping):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb["Individual Faculty TT"]
    assignments = []
    
    fac_headers = []
    for r in range(1, sheet.max_row + 1):
        val = get_cell_value(sheet, r, 1)
        if val and "faculty name" in str(val).lower():
            fac_headers.append((r, str(val)))
            
    for i, (r_header, header_text) in enumerate(fac_headers):
        fac_name = header_text.split("Faculty Name:")[-1].strip()
        fac_id = clean_id(fac_name)
        
        r_end = fac_headers[i+1][0] if i + 1 < len(fac_headers) else sheet.max_row
        
        day_rows = {}
        for r in range(r_header + 1, r_end):
            val = clean_text(sheet.cell(row=r, column=1).value).upper()
            if val in ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"]:
                day_rows[val] = r
                
        for day, r_day in day_rows.items():
            for p in range(1, 9):
                col = p + 1
                room_val = get_cell_value(sheet, r_day, col)
                subject_val = get_cell_value(sheet, r_day + 1, col)
                faculty_val = get_cell_value(sheet, r_day + 2, col)
                
                if room_val == "L" or not room_val:
                    room_val = "L"
                    
                parsed_slots = parse_grid_cell(room_val, subject_val, faculty_val)
                for slot in parsed_slots:
                    sub_raw = slot["subject"]
                    section_name = ""
                    # regex to isolate section name (e.g. CS-1, IT III, AIML II)
                    match = re.search(r'\b(cs|it|aiml|iot|ds|data science)\b\s*(-?\s*[1-3]|\biii\b|\bii\b|\bi\b)?', sub_raw.lower())
                    if match:
                        section_name = match.group(0).upper()
                        
                    assignments.append({
                        "faculty_id": fac_id,
                        "faculty_name": fac_name,
                        "faculty_initials": clean_text(faculty_val) or fac_id[:3].upper(),
                        "subject_code": slot["subject"],
                        "section_name": section_name or "Unknown",
                        "section_id": clean_id(section_name) if section_name else None,
                        "batch_label": slot["batch"],
                        "room_id": slot["room"],
                        "day": day.capitalize(),
                        "period": p,
                        "effective_from": "2024-09-30",
                        "effective_to": None,
                        "source": "faculty_wise"
                    })
                    
    return assignments

def parse_lab_wise_timetables(path, global_faculty_mapping):
    wb = openpyxl.load_workbook(path, data_only=True)
    assignments = []
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        lab_headers = []
        for r in range(1, sheet.max_row + 1):
            val = get_cell_value(sheet, r, 1)
            if val and "lab name" in str(val).lower():
                lab_headers.append((r, str(val)))
                
        for i, (r_header, header_text) in enumerate(lab_headers):
            lab_name = header_text.split("Lab Name-")[-1].strip()
            room_id = clean_id(lab_name).upper()
            
            eff_date = "2024-09-30"
            for c in range(1, sheet.max_column + 1):
                val = get_cell_value(sheet, r_header, c)
                if val and "effective from" in str(val).lower():
                    d = extract_date_from_string(str(val))
                    if d:
                        eff_date = d
                        break
                        
            r_end = lab_headers[i+1][0] if i + 1 < len(lab_headers) else sheet.max_row
            
            day_rows = {}
            for r in range(r_header + 1, r_end):
                val = clean_text(sheet.cell(row=r, column=1).value).upper()
                if val in ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"]:
                    day_rows[val] = r
                    
            for day, r_day in day_rows.items():
                for p in range(1, 9):
                    col = p + 1
                    
                    section_val = get_cell_value(sheet, r_day, col)
                    subject_val = get_cell_value(sheet, r_day + 1, col)
                    faculty_val = get_cell_value(sheet, r_day + 2, col)
                    
                    if not subject_val:
                        # Sometimes section and subject are swapped in lab sheets
                        subject_val = section_val
                        section_val = ""
                        
                    parsed_slots = parse_grid_cell(room_id, subject_val, faculty_val)
                    for slot in parsed_slots:
                        fac_initials = slot["faculty"].strip().upper()
                        fac_name = global_faculty_mapping.get(fac_initials, fac_initials)
                        fac_id = clean_id(fac_name) if fac_name else "unknown"
                        
                        s_val = clean_text(section_val)
                        s_name = ""
                        match = re.search(r'\b(cs|it|aiml|iot|ds|data science)\b\s*(-?\s*[1-3]|\biii\b|\bii\b|\bi\b)?', s_val.lower())
                        if match:
                            s_name = match.group(0).upper()
                        if not s_name:
                            # Try to parse section from subject cell
                            match2 = re.search(r'\b(cs|it|aiml|iot|ds)\s*(-?\s*[1-3]|\biii\b|\bii\b|\bi\b)?', slot["subject"].lower())
                            if match2:
                                s_name = match2.group(0).upper()
                                
                        assignments.append({
                            "faculty_id": fac_id,
                            "faculty_name": fac_name or fac_initials,
                            "faculty_initials": fac_initials,
                            "subject_code": slot["subject"],
                            "section_name": s_name or "Unknown",
                            "section_id": clean_id(s_name) if s_name else None,
                            "batch_label": slot["batch"],
                            "room_id": room_id,
                            "day": day.capitalize(),
                            "period": p,
                            "effective_from": eff_date,
                            "effective_to": None,
                            "source": "lab_wise"
                        })
                        
    return assignments

def seed_legacy_data(db: Session, excel_files: dict):
    fac_map = {}
    sub_map = {}
    
    for key in ["II Year_TT", "III Year_TT"]:
        path = excel_files.get(key)
        if path and os.path.exists(path):
            wb = openpyxl.load_workbook(path, data_only=True)
            f, s = parse_legends(wb)
            fac_map.update(f)
            sub_map.update(s)
            
    print(f"Constructed global maps: {len(fac_map)} faculty, {len(sub_map)} subjects.")
    
    # 2. Parse Load Distribution
    ld_path = excel_files.get("Load_Distribution")
    ld_records = []
    if ld_path and os.path.exists(ld_path):
        ld_records = parse_load_distribution(ld_path)
        
    db.query(LoadDistribution).delete()
    for rec in ld_records:
        entry = LoadDistribution(**rec)
        db.add(entry)
    db.commit()
    print(f"Imported {len(ld_records)} Load Distribution entries.")
    
    db.query(Faculty).delete()
    ld_faculties = set([r["faculty_name"] for r in ld_records if r["faculty_name"]])
    fac_instances = {}
    
    for full_name in ld_faculties:
        fac_id = clean_id(full_name)
        initials_list = []
        for initials, name in fac_map.items():
            if name.lower() in full_name.lower() or full_name.lower() in name.lower():
                initials_list.append(initials)
                
        dept = None
        for r in ld_records:
            if r["faculty_name"] == full_name:
                dept = r["section_name"]
                break
                
        fac = Faculty(
            id=fac_id,
            full_name=full_name,
            known_initials=",".join(initials_list) if initials_list else None,
            department=dept,
            max_weekly_hours=16
        )
        db.add(fac)
        fac_instances[fac_id] = fac
        
    for initials, full_name in fac_map.items():
        fac_id = clean_id(full_name)
        if fac_id not in fac_instances:
            fac = Faculty(
                id=fac_id,
                full_name=full_name,
                known_initials=initials,
                department=None,
                max_weekly_hours=16
            )
            db.add(fac)
            fac_instances[fac_id] = fac
            
    db.commit()
    print("Faculty seeded.")
    
    for fac_id, fac in fac_instances.items():
        user_exists = db.query(User).filter(User.username == fac_id).first()
        if not user_exists:
            user = User(
                username=fac_id,
                hashed_password=hash_password("faculty123"),
                role="faculty",
                faculty_id=fac_id
            )
            db.add(user)
    db.commit()
    print("Faculty user accounts seeded.")
    
    db.query(Subject).delete()
    subj_instances = {}
    
    for code, name in sub_map.items():
        stype = "lab" if "lab" in name.lower() or "practical" in name.lower() or "workshop" in name.lower() else "theory"
        sub_id = clean_id(code) or clean_id(name)
        
        if sub_id not in subj_instances:
            subject = Subject(
                id=sub_id,
                code=code,
                name=name,
                type=stype,
                weekly_hours=4 if stype == "theory" else 2,
                department=None
            )
            db.add(subject)
            subj_instances[sub_id] = subject
        
    for rec in ld_records:
        sub_name = rec["subject_name"]
        code = rec["semester"]
        sub_id = clean_id(code) or clean_id(sub_name)
        
        if sub_id not in subj_instances:
            stype = "lab" if rec["practical_hours"] > 0 else "theory"
            hours = int(rec["theory_hours"] + rec["practical_hours"])
            if hours == 0:
                hours = 4
                
            subject = Subject(
                id=sub_id,
                code=code or sub_name[:10],
                name=sub_name,
                type=stype,
                weekly_hours=hours,
                department=rec["section_name"]
            )
            db.add(subject)
            subj_instances[sub_id] = subject
            
    db.commit()
    print("Subjects seeded.")
    
    db.query(Batch).delete()
    db.query(Section).delete()
    
    for key, year in [("II Year_TT", 2), ("III Year_TT", 3)]:
        path = excel_files.get(key)
        if path and os.path.exists(path):
            wb_class = openpyxl.load_workbook(path, data_only=True)
            for sname in wb_class.sheetnames:
                sec_name = sname.strip()
                dept = "CS"
                if "it" in sec_name.lower():
                    dept = "IT"
                elif "aiml" in sec_name.lower():
                    dept = "AIML"
                elif "iot" in sec_name.lower():
                    dept = "IoT"
                elif "ds" in sec_name.lower():
                    dept = "DS"
                
                sec_id = f"{sec_name.replace(' ', '_')}_{year}"
                
                if not db.query(Section).filter(Section.id == sec_id).first():
                    sec = Section(
                        id=sec_id,
                        name=sec_name,
                        year=year,
                        department=dept
                    )
                    db.add(sec)
                    db.commit()
                    
                    for label in ["B1", "B2"]:
                        batch = Batch(
                            id=f"{sec_id}_{label}",
                            section_id=sec_id,
                            label=label
                        )
                        db.add(batch)
                    db.commit()
                    
    print("Sections and Batches seeded.")
    
    db.query(Room).delete()
    rooms_set = set()
    legacy_assignments = []
    
    ii_path = excel_files.get("II Year_TT")
    if ii_path and os.path.exists(ii_path):
        legacy_assignments.extend(parse_class_wise_timetables(ii_path, 2, fac_map))
        
    iii_path = excel_files.get("III Year_TT")
    if iii_path and os.path.exists(iii_path):
        legacy_assignments.extend(parse_class_wise_timetables(iii_path, 3, fac_map))
        
    lab_path = excel_files.get("Lab_Wise")
    if lab_path and os.path.exists(lab_path):
        legacy_assignments.extend(parse_lab_wise_timetables(lab_path, fac_map))
        
    fac_path = excel_files.get("Individual_Faculty")
    if fac_path and os.path.exists(fac_path):
        legacy_assignments.extend(parse_individual_faculty_wise(fac_path, fac_map))
        
    for a in legacy_assignments:
        room_name = a["room_id"].strip()
        if room_name and room_name.upper() not in ["L", "LUNCH", "LUNCH BREAK", "FREE", "LIBRARY", "SPORT", "SPORTS", "PDP", "APTTI", "APTT"]:
            clean_room = clean_id(room_name).upper()
            rooms_set.add(clean_room)
            
    for rname in rooms_set:
        rtype = "lab" if "lab" in rname.lower() or "center" in rname.lower() or "cc" in rname.lower() else "classroom"
        room = Room(
            id=rname,
            name=rname,
            type=rtype,
            capacity=60 if rtype == "classroom" else 30
        )
        db.add(room)
    db.commit()
    print(f"Rooms seeded: {list(rooms_set)}")
    
    db.query(Assignment).delete()
    
    seeded_count = 0
    for a in legacy_assignments:
        fac_id = a["faculty_id"]
        if not db.query(Faculty).filter(Faculty.id == fac_id).first():
            fac = Faculty(
                id=fac_id,
                full_name=a["faculty_name"],
                known_initials=a["faculty_initials"],
                department=None,
                max_weekly_hours=16
            )
            db.add(fac)
            db.commit()
            
        sub_code = a["subject_code"]
        sub_id = clean_id(sub_code)
        if not db.query(Subject).filter(Subject.id == sub_id).first():
            subject = Subject(
                id=sub_id,
                code=sub_code,
                name=sub_code,
                type="lab" if "lab" in sub_code.lower() or "pr" in sub_code.lower() else "theory",
                weekly_hours=4,
                department=None
            )
            db.add(subject)
            db.commit()
            
        room_id = a["room_id"].strip().upper()
        if not room_id or room_id in ["L", "LUNCH", "LUNCH BREAK", "FREE", "LIBRARY", "SPORT", "SPORTS", "PDP", "APTTI", "APTT"]:
            continue
            
        clean_room_id = clean_id(room_id).upper()
        if not db.query(Room).filter(Room.id == clean_room_id).first():
            room = Room(
                id=clean_room_id,
                name=room_id,
                type="lab" if "lab" in room_id.lower() or "center" in room_id.lower() else "classroom",
                capacity=60
            )
            db.add(room)
            db.commit()
            
        sec_id = a["section_id"]
        if not sec_id and a["section_name"]:
            s_name = a["section_name"].strip()
            sections = db.query(Section).all()
            for s in sections:
                if s.name.lower() in s_name.lower() or s_name.lower() in s.name.lower():
                    sec_id = s.id
                    break
                    
        day = a["day"].capitalize()
        period = a["period"]
        timeslot_id = f"{day}_{period}"
        
        if not db.query(TimeSlot).filter(TimeSlot.id == timeslot_id).first():
            continue
            
        batch_id = None
        if a["batch_label"] and sec_id:
            batch_id = f"{sec_id}_{a['batch_label']}"
            if not db.query(Batch).filter(Batch.id == batch_id).first():
                batch = Batch(
                    id=batch_id,
                    section_id=sec_id,
                    label=a["batch_label"]
                )
                db.add(batch)
                db.commit()
                
        assign = Assignment(
            faculty_id=fac_id,
            subject_id=sub_id,
            section_id=sec_id,
            batch_id=batch_id,
            room_id=clean_room_id,
            timeslot_id=timeslot_id,
            effective_from=a["effective_from"],
            effective_to=a["effective_to"],
            source=a["source"]
        )
        db.add(assign)
        seeded_count += 1
        
    db.commit()
    print(f"Seeded {seeded_count} legacy assignments.")
    return seeded_count

if __name__ == "__main__":
    from backend.database import SessionLocal, init_db
    init_db()
    db = SessionLocal()
    files = {
        "II Year_TT": r"C:\Users\tarun\Downloads\II Year_TT_July-Dec 24.xlsx",
        "III Year_TT": r"C:\Users\tarun\Downloads\III Year_TT_July-Dec 24 (2).xlsx",
        "Individual_Faculty": r"C:\Users\tarun\Downloads\Individual Faculty Wise Time Table July-Dec 2024 (1).xlsx",
        "Lab_Wise": r"C:\Users\tarun\Downloads\Lab Wise Time Table (1) (2).xlsx",
        "Load_Distribution": r"C:\Users\tarun\Downloads\Load Distribution_2024 (2).xlsx"
    }
    seed_legacy_data(db, files)
    db.close()
